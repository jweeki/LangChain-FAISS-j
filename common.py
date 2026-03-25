from pathlib import Path

import torch
from huggingface_hub import snapshot_download

from langchain_community.vectorstores import FAISS
from langchain_core.documents import Document
from langchain_huggingface import HuggingFaceEmbeddings

# TODO: 公共函数配置参数
BASE_DIR = Path(__file__).resolve().parent
MODEL_NAME = "shibing624/text2vec-base-chinese"
MODEL_DIR = BASE_DIR / "models" / "text2vec-base-chinese"
VECTOR_STORE_DIR = BASE_DIR / "faiss_store"


def get_device() -> str:
    return "cuda" if torch.cuda.is_available() else "cpu"


def get_storage_paths() -> dict[str, Path]:
    return {
        "model_dir": MODEL_DIR,
        "vector_store_dir": VECTOR_STORE_DIR,
    }


def download_embedding_model(
    model_name: str = MODEL_NAME, model_dir: Path = MODEL_DIR
) -> Path:
    model_dir.mkdir(parents=True, exist_ok=True)
    print(f"正在下载模型到: {model_dir}")
    snapshot_download(
        repo_id=model_name,
        local_dir=str(model_dir),
    )
    print("模型下载完成")
    return model_dir


def ensure_local_model(model_dir: Path = MODEL_DIR) -> Path:
    config_file = model_dir / "config.json"
    modules_file = model_dir / "modules.json"
    if config_file.exists() or modules_file.exists():
        return model_dir

    raise FileNotFoundError(
        f"未找到本地嵌入模型目录: {model_dir}\n"
        "请先下载模型，再执行存储或检索脚本。"
    )


def create_embeddings(model_dir: Path = MODEL_DIR) -> HuggingFaceEmbeddings:
    local_model_dir = ensure_local_model(model_dir)
    device = get_device()
    print(f"正在加载本地嵌入模型: {local_model_dir}")
    print(f"当前使用设备: {device}，加载中...")
    embeddings = HuggingFaceEmbeddings(
        model_name=str(local_model_dir),
        model_kwargs={
            "device": device,
            "local_files_only": True,
        },
        encode_kwargs={
            "batch_size": 32,
            "normalize_embeddings": True,
        },
        query_encode_kwargs={
            "batch_size": 32,
            "normalize_embeddings": True,
        },
        show_progress=False,
    )
    print("* 嵌入模型加载完成")
    return embeddings


def load_rows_from_excel(
    excel_path: str | Path,
) -> tuple[list[str], list[list[str]]]:
    """读取 Excel 所有列，返回 (列名列表, 行数据列表)。"""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ImportError(
            "读取 Excel 需要 openpyxl，请先在你的虚拟环境中安装它。"
        ) from exc

    excel_file = Path(excel_path)
    if not excel_file.exists():
        raise FileNotFoundError(f"未找到 Excel 文件: {excel_file}")

    workbook = load_workbook(excel_file, read_only=True, data_only=True)
    sheet = workbook.active

    rows = list(sheet.iter_rows(values_only=True))
    workbook.close()

    if not rows:
        raise ValueError(f"Excel 文件为空: {excel_file}")

    # 第一行作为列名
    raw_headers = [str(h).strip() if h is not None else "" for h in rows[0]]

    # 过滤掉全空列（列名为空且该列数据全为空）
    col_count = len(raw_headers)
    non_empty_cols = []
    for col_idx in range(col_count):
        if raw_headers[col_idx]:
            non_empty_cols.append(col_idx)
            continue
        # 列名为空，检查该列是否有数据
        has_data = any(
            row[col_idx] is not None and str(row[col_idx]).strip()
            for row in rows[1:]
            if col_idx < len(row)
        )
        if has_data:
            non_empty_cols.append(col_idx)

    headers = [raw_headers[i] if raw_headers[i] else f"col_{i}" for i in non_empty_cols]
    data = []
    for row in rows[1:]:
        values = [str(row[i]).strip() if i < len(row) and row[i] is not None else "" for i in non_empty_cols]
        if any(values):
            data.append(values)

    if not data:
        raise ValueError(f"Excel 没有可用数据行: {excel_file}")

    print(f"已从 Excel 读取 {len(data)} 行 x {len(headers)} 列: {excel_file}")
    print(f"列名: {headers}")
    return headers, data


def build_documents_from_rows(
    headers: list[str], rows: list[list[str]], search_col: int = 0
) -> list[Document]:
    """将多列数据构建为 Document，search_col 指定用于检索的列索引。"""
    documents = []
    for index, row in enumerate(rows, start=1):
        page_content = row[search_col] if search_col < len(row) else ""
        if not page_content:
            continue
        metadata = {"source": f"doc_{index}"}
        for i, header in enumerate(headers):
            if i != search_col and i < len(row):
                metadata[header] = row[i]
        documents.append(Document(page_content=page_content, metadata=metadata))
    return documents


def build_documents_from_excel(
    excel_path: str | Path, search_col: int = 0
) -> list[Document]:
    headers, rows = load_rows_from_excel(excel_path)
    print(f"检索列: [{search_col}] {headers[search_col]}")
    return build_documents_from_rows(headers, rows, search_col=search_col)


def build_vector_store(
    documents: list[Document], embeddings: HuggingFaceEmbeddings
) -> FAISS:
    print("正在构建 FAISS 向量库")
    vector_store = FAISS.from_documents(documents, embeddings)
    print("FAISS 向量库构建完成")
    return vector_store


def save_vector_store(vector_store: FAISS, store_dir: Path = VECTOR_STORE_DIR) -> None:
    store_dir.mkdir(parents=True, exist_ok=True)
    vector_store.save_local(folder_path=str(store_dir))
    print(f"向量库已保存到: {store_dir}")


def load_vector_store(
    embeddings: HuggingFaceEmbeddings, store_dir: Path = VECTOR_STORE_DIR
) -> FAISS:
    if not store_dir.exists():
        raise FileNotFoundError(
            f"未找到向量库目录: {store_dir}\n"
            "请先运行 store.py 生成本地索引。"
        )

    print(f"正在从以下目录加载向量库: {store_dir}")
    return FAISS.load_local(
        folder_path=str(store_dir),
        embeddings=embeddings,
        allow_dangerous_deserialization=True,
    )


def search_documents(
    vector_store: FAISS, query: str, k: int = 1
) -> list[tuple[Document, float]]:
    return vector_store.similarity_search_with_score(query, k=k)
